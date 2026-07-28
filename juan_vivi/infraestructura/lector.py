"""
Lector — infraestructura de entrada.
Lee archivos Excel AUTOFIN SP y los convierte en FilaEntrada del dominio.
Maneja limpieza básica de tipos (fechas como datetime, números, \xa0).
"""
import os
import openpyxl
from ..dominio.modelos import FilaEntrada


# ── Columnas esperadas (índice 0-based) ──────────────────────────
COL_NUMERO      = 0   # A — N°
COL_NOMBRE      = 1   # B — NOMBRE
COL_RUT         = 2   # C — RUT
COL_NOMBRE_PARA = 3   # D — NOMBRE PARA
COL_RUT_PARA    = 4   # E — RUT COMPRA PARA
COL_OPERACION   = 5   # F — OPERACIÓN
COL_PATENTE     = 6   # G — PATENTE
COL_TIPO        = 10  # K — COMUNIDAD | COMPRA PARA | None

COLUMNAS_REQUERIDAS = 11   # mínimo de columnas que debe tener la hoja


def _str_o_none(valor) -> str | None:
    """Convierte cualquier valor de celda a str limpio, o None si está vacío."""
    if valor is None:
        return None
    texto = str(valor).replace("\xa0", "").strip()
    return texto if texto else None


def _str_requerido(valor) -> str:
    """Como _str_o_none pero retorna '' en vez de None."""
    resultado = _str_o_none(valor)
    return resultado or ""


def leer_archivo(ruta: str) -> list:
    """
    Lee un Excel AUTOFIN SP y retorna lista[FilaEntrada].
    Lanza ValueError con mensaje descriptivo si el formato no es válido.
    """
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"No se encontró el archivo: {ruta}")

    nombre_archivo = os.path.basename(ruta)

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo abrir '{nombre_archivo}': {e}")

    ws = wb.active

    # Validar que tenga suficientes columnas
    if ws.max_column < COLUMNAS_REQUERIDAS:
        raise ValueError(
            f"'{nombre_archivo}' tiene solo {ws.max_column} columnas — "
            f"se esperan al menos {COLUMNAS_REQUERIDAS}."
        )

    filas: list[FilaEntrada] = []

    # La fila 1 es encabezado — empezar desde fila 2
    for num_fila, fila_excel in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Ignorar filas completamente vacías al final del archivo
        if all(c is None for c in fila_excel):
            continue

        entrada = FilaEntrada(
            numero=fila_excel[COL_NUMERO] if fila_excel[COL_NUMERO] is not None else 0,
            nombre=_str_requerido(fila_excel[COL_NOMBRE]),
            rut=_str_requerido(fila_excel[COL_RUT]),
            nombre_para=_str_o_none(fila_excel[COL_NOMBRE_PARA]),
            rut_para=_str_o_none(fila_excel[COL_RUT_PARA]),
            operacion=_str_o_none(fila_excel[COL_OPERACION]),
            patente=_str_o_none(fila_excel[COL_PATENTE]),
            tipo=_str_o_none(fila_excel[COL_TIPO]),
            num_fila_excel=num_fila,
        )
        filas.append(entrada)

    return filas


def leer_multiples(rutas: list) -> tuple[list, list]:
    """
    Lee varios archivos Excel y consolida las filas.
    Retorna (filas_consolidadas, errores_por_archivo).
    Los errores no detienen el proceso — se reportan y se sigue con el siguiente.
    """
    todas_las_filas: list[FilaEntrada] = []
    errores: list[str] = []

    for ruta in rutas:
        try:
            filas = leer_archivo(ruta)
            todas_las_filas.extend(filas)
        except (FileNotFoundError, ValueError) as e:
            errores.append(str(e))

    return todas_las_filas, errores
