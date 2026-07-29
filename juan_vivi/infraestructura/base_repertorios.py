"""
Persistencia del historial de repertorios de la 33ª Notaría.

Soporta múltiples formatos de planilla (uno por función cargar_desde_*).
Todos comparten la misma función base agregar_registro().

Ruta cuando está instalado : %LOCALAPPDATA%\\JUAN-VIVI\\repertorios.json
Ruta en desarrollo          : data/repertorios.json (raíz del proyecto)

Los repertorios son únicos — si ya existe uno al cargar, se reporta
como duplicado sin sobreescribir. El usuario decide si reemplaza.
"""

import json
import os
import sys
from pathlib import Path
from datetime import datetime


# ── Ruta del archivo ──────────────────────────────────────────────────────────

def _ruta_datos() -> Path:
    if getattr(sys, "frozen", False):
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
        return base / "JUAN-VIVI" / "repertorios.json"
    return Path(__file__).parent.parent.parent / "data" / "repertorios.json"

_RUTA_JSON = _ruta_datos()


# ── Persistencia base ─────────────────────────────────────────────────────────

def _cargar() -> dict:
    if _RUTA_JSON.exists():
        return json.loads(_RUTA_JSON.read_text(encoding="utf-8"))
    return {}


def _guardar(registros: dict) -> None:
    _RUTA_JSON.parent.mkdir(parents=True, exist_ok=True)
    _RUTA_JSON.write_text(
        json.dumps(registros, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── API pública ───────────────────────────────────────────────────────────────

def buscar(numero: str) -> dict | None:
    """Retorna el registro del repertorio o None si no existe."""
    return _cargar().get(str(numero).strip())


def reemplazar_registro(numero: str, datos: dict) -> None:
    """Sobreescribe un registro existente (usado al resolver un duplicado)."""
    registros = _cargar()
    registros[str(numero).strip()] = datos
    _guardar(registros)


def agregar_registro(numero: str, datos: dict, registros: dict) -> bool:
    """
    Intenta agregar un registro al dict en memoria.
    Retorna True si es nuevo, False si ya existía (duplicado).
    No persiste — el llamador llama a _guardar() al final.
    """
    clave = str(numero).strip()
    if clave in registros:
        return False
    registros[clave] = datos
    return True


# ── Formato 1: planilla "juan REPERTORIOS" ────────────────────────────────────

def _extraer_compareciente(texto: str) -> str:
    """
    El campo Compareciente tiene dos líneas: la empresa y la persona.
    Filtra la línea de la empresa (AUTOFIN / GLOBAL) y retorna la persona.
    """
    if not texto:
        return ""
    lineas = [l.strip() for l in str(texto).replace("\r\n", "\n").split("\n") if l.strip()]
    excluir = ("autofin", "global soluciones")
    persona = [l for l in lineas if not any(e in l.lower() for e in excluir)]
    return persona[0] if persona else lineas[-1] if lineas else ""


def cargar_desde_juan_repertorios(ruta: str) -> tuple[int, list[dict], list[str]]:
    """
    Lee la planilla "juan REPERTORIOS" (hoja Exportar).

    Columnas usadas:
      A (0) OT  |  C (2) Repertorio  |  D (3) Fecha Repertorio
      F (5) Cliente  |  H (7) Materia  |  I (8) Compareciente

    Retorna (cargados, duplicados, errores).
    - cargados   : cantidad de registros nuevos guardados
    - duplicados : lista de {numero, existente, entrante} para que el usuario decida
    - errores    : lista de strings con filas que no se pudieron procesar
    """
    import openpyxl

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb["Exportar"] if "Exportar" in wb.sheetnames else wb.active

    registros   = _cargar()
    cargados    = 0
    duplicados: list[dict] = []
    errores:    list[str]  = []

    for fila_idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rep_raw = fila[2]   # col C — Repertorio
        if not rep_raw:
            continue
        try:
            numero = str(rep_raw).strip()

            # Fecha
            fecha_raw = fila[3]
            if isinstance(fecha_raw, datetime):
                fecha = fecha_raw.strftime("%Y-%m-%d")
            else:
                fecha = str(fecha_raw).split(" ")[0] if fecha_raw else ""

            datos = {
                "repertorio":     numero,
                "ot":             str(int(fila[0])) if fila[0] else "",
                "fecha":          fecha,
                "cliente":        str(fila[5] or "").strip(),
                "materia":        str(fila[7] or "").strip(),
                "compareciente":  _extraer_compareciente(fila[8]),
            }

            es_nuevo = agregar_registro(numero, datos, registros)
            if es_nuevo:
                cargados += 1
            else:
                duplicados.append({
                    "numero":    numero,
                    "existente": registros[numero],
                    "entrante":  datos,
                })
        except Exception as e:
            errores.append(f"Fila {fila_idx}: {e}")

    _guardar(registros)
    return cargados, duplicados, errores
