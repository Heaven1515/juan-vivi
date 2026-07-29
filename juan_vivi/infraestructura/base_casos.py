"""
Persistencia de casos de firma en archivo JSON.

Reemplaza casos_prueba.py — los datos sobreviven entre sesiones.
Cuando el sidecar está empaquetado (frozen), el archivo se guarda en
%LOCALAPPDATA%\JUAN-VIVI\casos.json para que persista entre versiones.
En desarrollo se guarda en data/casos.json en la raíz del proyecto.
Los 3 casos de prueba originales se insertan si el archivo no existe.
"""

import json
import os
import sys
from pathlib import Path

def _ruta_datos() -> Path:
    if getattr(sys, "frozen", False):
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
        return base / "JUAN-VIVI" / "casos.json"
    return Path(__file__).parent.parent.parent / "data" / "casos.json"

_RUTA_JSON = _ruta_datos()

def _cargar() -> dict:
    if _RUTA_JSON.exists():
        return json.loads(_RUTA_JSON.read_text(encoding="utf-8"))
    return {}


def _guardar(casos: dict) -> None:
    _RUTA_JSON.parent.mkdir(parents=True, exist_ok=True)
    _RUTA_JSON.write_text(
        json.dumps(casos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def buscar_caso(numero: str) -> dict | None:
    """Retorna el caso para el número de repertorio, o None si no existe."""
    return _cargar().get(str(numero).strip())


def agregar_caso(numero: str, anho: str, tipo_contrato: str,
                 fecha_dia: int, fecha_mes: int, fecha_anio: int) -> None:
    """Agrega o sobreescribe un caso en la base de datos JSON."""
    casos = _cargar()
    casos[str(numero).strip()] = {
        "numero":        str(numero).strip(),
        "anho":          str(anho).strip(),
        "tipo_contrato": str(tipo_contrato).strip().upper(),
        "fecha_dia":     int(fecha_dia),
        "fecha_mes":     int(fecha_mes),
        "fecha_anio":    int(fecha_anio),
    }
    _guardar(casos)


def eliminar_caso(numero: str) -> bool:
    """Elimina un caso. Retorna True si existía."""
    casos = _cargar()
    if str(numero).strip() in casos:
        del casos[str(numero).strip()]
        _guardar(casos)
        return True
    return False


def listar_casos() -> list[dict]:
    """Retorna todos los casos ordenados por número."""
    casos = _cargar()
    return sorted(casos.values(), key=lambda c: int(c["numero"]) if c["numero"].isdigit() else 0)


def cargar_desde_excel(ruta: str) -> tuple[int, list[str]]:
    """
    Lee un Excel con columnas:
      N° Repertorio | Año | Tipo de Contrato | Día | Mes | Año escritura

    Detecta las columnas por su encabezado (fila 1) buscando palabras clave.
    Agrega o sobreescribe casos en la BD.
    Retorna (cantidad_cargada, lista_de_errores).
    """
    import openpyxl

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb.active

    # ── Detectar columnas por encabezado (fila 1) ─────────────────────────
    encabezados = [str(c.value or "").strip().lower() for c in ws[1]]

    def _col(palabras: list[str]) -> int | None:
        for i, h in enumerate(encabezados):
            if any(p in h for p in palabras):
                return i
        return None

    ci_num   = _col(["repertorio", "n°", "numero", "número", "rep"])
    ci_anho  = _col(["año rep", "anho rep", "año_r", "año r"])
    if ci_anho is None:
        ci_anho = _col(["año", "anho", "year"])
    ci_tipo  = _col(["tipo", "materia", "contrato"])
    ci_dia   = _col(["día", "dia"])
    ci_mes   = _col(["mes"])
    ci_anio  = _col(["año escr", "anho escr", "año_e", "año e", "año esc"])
    if ci_anio is None and ci_anho is not None:
        # Buscar una segunda columna con "año" (distinta de ci_anho)
        for i, h in enumerate(encabezados):
            if "año" in h or "anho" in h:
                if i != ci_anho:
                    ci_anio = i
                    break

    faltantes = []
    for nombre, idx in [("N° Repertorio", ci_num), ("Tipo de Contrato", ci_tipo),
                        ("Día", ci_dia), ("Mes", ci_mes)]:
        if idx is None:
            faltantes.append(nombre)
    if faltantes:
        raise ValueError(f"Columnas no encontradas: {', '.join(faltantes)}")

    casos = _cargar()
    cargados = 0
    errores: list[str] = []

    for fila_idx, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        num_raw = fila[ci_num] if ci_num is not None else None
        if not num_raw:
            continue  # fila vacía
        try:
            numero = str(int(float(str(num_raw).strip())))
            anho   = str(int(float(str(fila[ci_anho]).strip()))) if ci_anho is not None and fila[ci_anho] else "2026"
            tipo   = str(fila[ci_tipo] or "").strip().upper() if ci_tipo is not None else ""
            dia    = int(float(str(fila[ci_dia]).strip())) if ci_dia is not None and fila[ci_dia] else 1
            mes    = int(float(str(fila[ci_mes]).strip())) if ci_mes is not None and fila[ci_mes] else 1
            anio_e = int(float(str(fila[ci_anio]).strip())) if ci_anio is not None and fila[ci_anio] else int(anho)

            casos[numero] = {
                "numero": numero, "anho": anho, "tipo_contrato": tipo,
                "fecha_dia": dia, "fecha_mes": mes, "fecha_anio": anio_e,
            }
            cargados += 1
        except Exception as e:
            errores.append(f"Fila {fila_idx}: {e}")

    _guardar(casos)
    return cargados, errores
