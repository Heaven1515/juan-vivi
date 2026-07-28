"""
Escritor — infraestructura de salida.
Genera el Excel de planilla masivos replicando el formato del archivo de referencia:
  - Fila 1: encabezados de grupo combinados (A1:D1, E1:H1, I1:T1)
  - Fila 2: encabezados de columnas individuales, altura 30pt, borde inferior thin
  - Filas 3+: datos
  - Anchos de columna según el archivo original
  - Fuente Calibri 11, sin negritas (replicando el original)
"""
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..dominio.modelos import FilaSalida


# ── Anchos de columna (extraídos del archivo de referencia) ──────
ANCHOS_COLUMNA = {
    1:  13.0,       # A — RUT c1
    2:  41.14,      # B — Apellido Paterno c1
    3:  20.57,      # C — Apellido Materno c1
    4:  33.57,      # D — Nombres c1
    5:  13.0,       # E — RUT c2
    6:  41.14,      # F — Apellido Paterno c2
    7:  20.57,      # G — Apellido Materno c2
    8:  33.57,      # H — Nombres c2
    9:  13.86,      # I — Tasación Fiscal
    10: 15.0,       # J — Precio de Venta
    11: 12.86,      # K — Patente
    12: 8.0,        # L — Tipo
    13: 10.0,       # M — Marca
    14: 14.0,       # N — Modelo
    15: 6.0,        # O — Año
    16: 8.0,        # P — Color
    17: 10.0,       # Q — Motor
    18: 12.0,       # R — Chasis
    19: 10.0,       # S — Serie
    20: 10.0,       # T — Vin
    21: 19.57,      # U — Código de Operación
}

# ── Encabezados ───────────────────────────────────────────────────
GRUPOS_F1 = [
    ("Compareciente 1 / Vendedor",  "A1", "D1"),
    ("Compareciente 2 / Comprador", "E1", "H1"),
    ("Vehículo",                    "I1", "T1"),
    # U1 va solo (sin combinar)
]

ENCABEZADOS_F2 = [
    "RUT",
    "Apellido Paterno /\nRazón Social",
    "Apellido Materno",
    "Nombres",
    "RUT",
    "Apellido Paterno /\nRazón Social",
    "Apellido Materno",
    "Nombres",
    "Tasación Fiscal",
    "Precio de Venta",
    "Patente",
    "Tipo",
    "Marca",
    "Modelo",
    "Año",
    "Color",
    "Motor",
    "Chasis",
    "Serie",
    "Vin",
    "Código de Operación",
]


# ── Helpers de estilo ─────────────────────────────────────────────

def _fuente(bold: bool = False) -> Font:
    return Font(name="Calibri", size=11, bold=bold)

def _borde_inferior() -> Border:
    delgado = Side(border_style="thin")
    return Border(bottom=delgado, top=delgado)

def _centrado() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _izquierda() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


# ── Función principal ─────────────────────────────────────────────

def escribir_planilla(filas_salida: list, ruta_destino: str) -> None:
    """
    Genera el Excel de planilla masivos en ruta_destino.
    Recibe lista[FilaSalida] ya transformadas.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # ── Anchos de columna ────────────────────────────────────────
    for num_col, ancho in ANCHOS_COLUMNA.items():
        ws.column_dimensions[get_column_letter(num_col)].width = ancho

    # ── Fila 1: encabezados de grupo combinados ───────────────────
    for texto, inicio, fin in GRUPOS_F1:
        ws.merge_cells(f"{inicio}:{fin}")
        celda = ws[inicio]
        celda.value     = texto
        celda.font      = _fuente()
        celda.alignment = _centrado()

    # U1 — sin combinar
    ws["U1"].value     = "Datos OT"
    ws["U1"].font      = _fuente()
    ws["U1"].alignment = _izquierda()

    ws.row_dimensions[1].height = 18

    # ── Fila 2: encabezados individuales ─────────────────────────
    for col_idx, titulo in enumerate(ENCABEZADOS_F2, start=1):
        celda = ws.cell(row=2, column=col_idx)
        celda.value     = titulo
        celda.font      = _fuente()
        celda.alignment = _centrado()
        celda.border    = _borde_inferior()

    ws.row_dimensions[2].height = 30

    # ── Filas de datos ────────────────────────────────────────────
    for fila_idx, fila in enumerate(filas_salida, start=3):
        valores = [
            # Compareciente 1
            fila.c1_rut,
            fila.c1_apellido_paterno,
            fila.c1_apellido_materno,
            fila.c1_nombres,
            # Compareciente 2
            fila.c2_rut,
            fila.c2_apellido_paterno,
            fila.c2_apellido_materno,
            fila.c2_nombres,
            # Vehículo
            fila.tasacion_fiscal,
            fila.precio_venta,
            fila.patente,
            fila.tipo_vehiculo,
            fila.marca,
            fila.modelo,
            fila.anio,
            fila.color,
            fila.motor,
            fila.chasis,
            fila.serie,
            fila.vin,
            # Datos OT
            fila.codigo_operacion,
        ]
        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila_idx, column=col_idx, value=valor)
            celda.font      = _fuente()
            celda.alignment = _izquierda()

    # ── Guardar ───────────────────────────────────────────────────
    os.makedirs(os.path.dirname(ruta_destino) or ".", exist_ok=True)
    wb.save(ruta_destino)


def nombre_archivo_salida(carpeta_destino: str = ".", prefijo_sp: str = "") -> str:
    """
    Genera el nombre del archivo de salida.
    Si se entrega prefijo_sp (ej. "3206"), el nombre será:
      Repertorios Masivo SP 3206.xlsx
    Si son varios archivos:
      Repertorios Masivo SP 3206_3217.xlsx
    Sin prefijo:
      planilla_masivos_2026-07-27_09-14.xlsx
    """
    if prefijo_sp:
        nombre = f"Repertorios Masivo SP {prefijo_sp}.xlsx"
    else:
        marca = datetime.now().strftime("%Y-%m-%d_%H-%M")
        nombre = f"planilla_masivos_{marca}.xlsx"
    return os.path.join(carpeta_destino, nombre)
